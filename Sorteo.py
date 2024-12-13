import config
import subprocess
import sys
import random
import time
from datetime import datetime

def verificar_e_instalar_libreria(nombre_libreria):
    try:
        # Intentar importar la librería
        __import__(nombre_libreria)
        # print(f"La librería '{nombre_libreria}' ya está instalada.")
    except ImportError:
        print(f"La librería '{nombre_libreria}' no está instalada. Procediendo a instalarla...")
        # Intentar instalar la librería con pip
        subprocess.check_call([sys.executable, "-m", "pip", "install", nombre_libreria])
        print(f"La librería '{nombre_libreria}' ha sido instalada exitosamente.")

# Verificar e instalar 'openpyxl' si es necesario
verificar_e_instalar_libreria("openpyxl")

# Importar 'openpyxl' después de verificar/instalar
import openpyxl

# Cargar o crear el archivo Excel
def cargar_o_crear_excel(nombre_archivo):
    try:
        # Intentar cargar el archivo existente
        workbook = openpyxl.load_workbook(nombre_archivo)
        print(workbook.sheetnames)
        
        # Verificar si las pestañas requeridas existen
        if config.PesParticipante not in workbook.sheetnames or config.PesHistoria not in workbook.sheetnames:
            print("El archivo existe, pero faltan pestañas requeridas. Se agregarán.")
            if config.PesParticipante not in workbook.sheetnames:
                participante = workbook.create_sheet(config.PesParticipante)
                participante.append([config.ColNombre])  # Encabezado
            if config.PesHistoria not in workbook.sheetnames:
                historia = workbook.create_sheet(config.PesHistoria)
                historia.append([config.ColAnio, config.ColParticipante, config.ColAmigo])  # Encabezado
            try:
                workbook.save(nombre_archivo)
            except PermissionError:
                print("\nError: Asegúrese de que el archivo Excel esté cerrado antes de intentar guardar.")
            except Exception as e:
                print(f"\nError inesperado al guardar los resultados: {e}")
    
    except FileNotFoundError:
        # Crear un nuevo archivo Excel con las pestañas y encabezados necesarios
        workbook = openpyxl.Workbook()
        
        # Crear y configurar las pestañas
        historia = workbook.active
        historia.title = config.PesHistoria
        historia.append([config.ColAnio, config.ColParticipante, config.ColAmigo])  # Encabezados
        
        participante = workbook.create_sheet(config.PesParticipante)
        participante.append([config.ColNombre])  # Encabezado
        
        # Guardar el nuevo archivo
        try:
            workbook.save(nombre_archivo)
        except PermissionError:
            print("\nError: Asegúrese de que el archivo Excel esté cerrado antes de intentar guardar.")
        except Exception as e:
            print(f"\nError inesperado al guardar los resultados: {e}")
        
        print(f"Archivo '{nombre_archivo}' creado con las pestañas necesarias.")
    
    return workbook

# Método para agregar o editar participantes
def agregar_o_editar_participantes(workbook, nombre_archivo):
    # Obtener la hoja "Participante"
    if config.PesParticipante not in workbook.sheetnames:
        print("La pestaña 'Participante' no existe. Creándola...")
        participante_sheet = workbook.create_sheet(config.PesParticipante)
        participante_sheet.append([config.ColNombre])  # Agregar encabezado
    else:
        participante_sheet = workbook[config.PesParticipante]

    # Verificar si ya hay participantes
    participantes_existentes = list(participante_sheet.iter_rows(min_row=2, values_only=True))
    if participantes_existentes:
        decision = input("Se han encontrado participantes existentes. ¿Deseas sobrescribirlos? (s/n): ").strip().lower()
        if decision != "s":
            print("No se realizaron cambios. Regresando al menú principal.")
            return
        
        # Borrar completamente las filas existentes (manteniendo el encabezado)
        participante_sheet.delete_rows(2, participante_sheet.max_row - 1)

    # Solicitar la lista de participantes
    print("\nIngresa una lista con el siguiente formato: 'amigo_1, amigo_2, ..., amigo_n (Al menos 3 amigos)'")
    input_participantes = input("Participantes: ").strip()

    # Validar el formato de la lista de participantes
    participantes = [p.strip() for p in input_participantes.split(",") if p.strip()]
    if len(participantes) < 3:
        print("Error: Debes ingresar al menos 3 participantes.")
        return

    if len(participantes) != len(input_participantes.split(",")):
        print("Error: El formato ingresado es incorrecto. Asegúrate de separar los nombres con comas.")
        return

    # Agregar los nuevos participantes
    for participante in participantes:
        participante_sheet.append([participante])

    # Guardar los cambios en el archivo
    try:
        workbook.save(nombre_archivo)
    except PermissionError:
        print("\nError: Asegúrese de que el archivo Excel esté cerrado antes de intentar guardar.")
    except Exception as e:
        print(f"\nError inesperado al guardar los resultados: {e}")

    print("Participantes actualizados correctamente en la base de datos.")

# Método para ver sorteos de un año específico
def ver_sorteos_por_anio(workbook, anio, anio_actual):
    historia = workbook[config.PesHistoria]
    filas_encontradas = [
        fila for fila in historia.iter_rows(min_row=2, values_only=True) if fila[0] == anio
    ]
    
    if filas_encontradas:
        if anio == anio_actual:
            print(f"\nEste año los amigos secretos son:")
            print(f"==========")
            for fila in filas_encontradas:
                print(f"{fila[1]}, su amigo secreto es {fila[2]}")
        else:
            print(f"\nEl {anio} los amigos secretos fueron:")
            print(f"==========")
            for fila in filas_encontradas:
                print(f"{fila[1]}, su amigo secreto fue {fila[2]}")
        
        # Mostrar ciclos formados
        print("\nCiclos formados en el sorteo:")
        participantes = [fila[1] for fila in filas_encontradas]
        amigos = [fila[2] for fila in filas_encontradas]
        
        visitados = set()
        for participante in participantes:
            if participante not in visitados:
                ciclo = []
                actual = participante
                while actual not in ciclo:
                    ciclo.append(actual)
                    visitados.add(actual)
                    actual = amigos[participantes.index(actual)]  # Obtener el amigo secreto
                print(" -> ".join(ciclo) + f" -> {ciclo[0]}")  # Ciclo completo
        return filas_encontradas
    else:
        print(f"\nNo se encontraron registros de sorteos para el año {anio}.")
        return []

def listar_participantes(workbook):
    """ 
        Lee todos los nombres de la pestaña 'Participante' en la base de datos.
        Args:
        workbook: Objeto Workbook cargado con openpyxl.
        Returns:
        Una lista de nombres encontrados en la columna 'Nombre'.
    """
    participante = workbook[config.PesParticipante]
    # Leer todos los valores en la columna 'Nombre', omitiendo el encabezado
    nombres = [fila[0] for fila in participante.iter_rows(min_row=2, max_col=1, values_only=True) if fila[0]]
    
    if nombres:
        print("\nLista de participantes:")
        for i, nombre in enumerate(nombres, start=1):
            print(f"{i}. {nombre}")
    else:
        print("\nNo se encontraron participantes en la base de datos.")
    
    return nombres

def preguntar_por_participante(workbook):
    """
    Pregunta por un participante y muestra su historial de amigos secretos,
    excluyendo el presente año.
    """
    historia_hoja = workbook[config.PesHistoria]

    participante = input("Ingrese el nombre del participante: ").strip().lower()

    registros = [
        (fila[0], fila[1], fila[2])  # Año, Participante, Amigo
        for fila in historia_hoja.iter_rows(min_row=2, values_only=True)
    ]

    resultados = [
        (anio, amigo)
        for anio, part, amigo in registros
        if part and amigo and part.lower() == participante
    ]

    if resultados:
        print(f"\nPara la persona {participante.capitalize()}:")
        for anio, amigo in resultados:
            print(f"En el año {anio} su amigo secreto es/fue {amigo}.")
        return resultados
    else:
        print(f"\nNo se encontraron registros para {participante.capitalize()}.")
        return []

def realizar_sorteo(workbook, anio_actual, nombre_archivo):
    """
    Realiza un sorteo de amigo secreto cumpliendo con las restricciones indicadas.
    Simula un cargador antes de los resultados, muestra los ciclos formados y valida si el año ya existe en la base de datos antes de guardar.

    Args:
        workbook: Objeto Workbook de openpyxl.
        anio_actual: Año en el que se realiza el sorteo.

    Returns:
        Lista de pares (participante, amigo) indicando las asignaciones del sorteo si se confirman.
    """
    participantes_hoja = workbook[config.PesParticipante]  # Cambiado a "Participante"
    historia_hoja = workbook[config.PesHistoria]

    # Leer lista de participantes
    participantes = [
        fila[0] for fila in participantes_hoja.iter_rows(min_row=2, max_col=1, values_only=True) if fila[0]
    ]

    if len(participantes) < 2:
        print("Error: Se necesitan al menos dos participantes para realizar el sorteo.")
        return []

    def obtener_historial(anio_inicio, anio_fin):
        """Obtiene historial de regalos entre los años indicados."""
        return [
            (fila[1], fila[2])  # (Participante, Amigo)
            for fila in historia_hoja.iter_rows(min_row=2, values_only=True)
            if anio_inicio <= fila[0] <= anio_fin
        ]

    # Validar el historial de los últimos 5 años
    historial_reciente = obtener_historial(anio_actual - 5, anio_actual - 1)

    # Generar asignaciones válidas
    while True:
        asignaciones = participantes[:]
        random.shuffle(asignaciones)  # Mezclar los participantes

        es_valido = True
        ciclos = {}
        for i, participante in enumerate(participantes):
            amigo = asignaciones[i]
            if participante == amigo:
                es_valido = False  # No puede regalarse a sí mismo
                break

            # Verificar si hay ciclos pequeños (A->B->A o ciclos de 3)
            if amigo in ciclos:
                if ciclos[amigo] == participante:  # Detectar ciclos de 2
                    es_valido = False
                    break
                if len(ciclos) == 2 and amigo in ciclos.values():  # Detectar ciclos de 3
                    es_valido = False
                    break

            ciclos[participante] = amigo

            # Validar que el historial reciente no tenga repeticiones
            if (participante, amigo) in historial_reciente:
                es_valido = False
                break

        if es_valido:
            # Simulación de cargador
            print("\nRealizando sorteo, por favor espera...")
            duracion_carga = random.randint(2, 5)  # Duración aleatoria entre 2 y 5 segundos
            for i in range(duracion_carga):
                print(f"\rCargando{'.' * ((i % 3) + 1)}{' ' * (3 - ((i % 3) + 1))}", end="")
                time.sleep(1)
            print("\r¡Sorteo completado!                 ")

            # Mostrar resultados en formato original
            print(f"\nSorteo realizado para el año {anio_actual}:")
            for participante, amigo in zip(participantes, asignaciones):
                print(f"{participante} regala a {amigo}")

            # Mostrar ciclos formados
            print("\nCiclos formados en el sorteo:")
            visitados = set()
            for participante in participantes:
                if participante not in visitados:
                    ciclo = []
                    actual = participante
                    while actual not in ciclo:
                        ciclo.append(actual)
                        visitados.add(actual)
                        actual = asignaciones[participantes.index(actual)]
                    print(" -> ".join(ciclo) + f" -> {ciclo[0]}")  # Ciclo completo

            # Verificar si el año ya existe en la base de datos
            anio_existente = [
                fila[0] for fila in historia_hoja.iter_rows(min_row=2, max_col=1, values_only=True)
            ]
            if anio_actual in anio_existente:
                print(f"\nAdvertencia: Ya existe información para el año {anio_actual}.")
                confirmar_sobrescribir = input(
                    "¿Deseas sobrescribir los registros existentes? (sí/no): "
                ).strip().lower()
                if confirmar_sobrescribir not in ["s", "si"]:
                    print("El sorteo no fue guardado. Regresando al menú principal.")
                    return []

                # Eliminar registros del año actual
                filas_a_eliminar = [
                    idx for idx, fila in enumerate(historia_hoja.iter_rows(min_row=2, values_only=True), start=2)
                    if fila[0] == anio_actual
                ]
                for idx in reversed(filas_a_eliminar):
                    historia_hoja.delete_rows(idx)

            # Confirmación antes de guardar
            confirmar = input("\n¿Confirmas los resultados para guardar? (sí/no): ").strip().lower()
            if confirmar in ["s", "si"]:
                # Guardar resultados en la pestaña config.PesHistoria
                try:
                    for participante, amigo in zip(participantes, asignaciones):
                        historia_hoja.append([anio_actual, participante, amigo])
                    workbook.save(nombre_archivo)
                    print("Resultados guardados exitosamente.")
                except PermissionError:
                    print("\nError: Asegúrese de que el archivo Excel esté cerrado antes de intentar guardar.")
                except Exception as e:
                    print(f"\nError inesperado al guardar los resultados: {e}")
                return list(zip(participantes, asignaciones))
            else:
                print("El sorteo no fue guardado. Regresando al menú principal.")
                return []

# Verificar e instalar 'openpyxl' si es necesario
verificar_e_instalar_libreria("PIL")
verificar_e_instalar_libreria("os")

from PIL import Image, ImageDraw, ImageFont
import os

def seleccionar_frase_unica(frases_graciosas, frases_usadas):
    """
    Selecciona una frase aleatoria que no haya sido seleccionada previamente.
    Una vez que todas las frases han sido seleccionadas, reinicia la lista de frases usadas.
    
    Args:
        frases_graciosas: Lista de frases posibles.
        frases_usadas: Lista de frases ya seleccionadas.
    
    Returns:
        str: La frase seleccionada aleatoriamente.
    """
    if len(frases_usadas) == len(frases_graciosas):
        # Si ya se han usado todas las frases, reiniciar la lista de frases usadas
        frases_usadas.clear()
    # Seleccionar una frase aleatoria que no esté en frases_usadas
    frase = random.choice([f for f in frases_graciosas if f not in frases_usadas])
    # Marcar la frase como usada
    frases_usadas.append(frase)
    return frase

def ilustrar_amigos_por_anio(workbook, anio_seleccionado, fondo_imagen):
    """
    Genera imágenes personalizadas para los amigos secretos de un año seleccionado,
    basado en la base de datos de la pestaña 'Historia'.
    """
    historia_hoja = workbook[config.PesHistoria]
    carpeta_destino = "Difusion"

    # Frases graciosas
    frases_graciosas = config.frases
    frases_usadas = []

    # Validar existencia de la imagen base
    if not os.path.exists(fondo_imagen):
        print(f"No se encontró el archivo {fondo_imagen}. Por favor, asegúrate de que esté en la carpeta.")
        return

    # Recuperar datos del año seleccionado
    registros = [
        (fila[1], fila[2])  # Participante, Amigo
        for fila in historia_hoja.iter_rows(min_row=2, values_only=True)
        if fila[0] == anio_seleccionado
    ]

    if not registros:
        print(f"No hay datos de amigos secretos para el año {anio_seleccionado}.")
        return

    # Verificar existencia de la carpeta
    if os.path.exists(carpeta_destino):
        confirmar = input(f"La carpeta {carpeta_destino} ya existe. ¿Deseas sobrescribirla? (s/n): ").strip().lower()
        if confirmar not in ["s", "si"]:
            print("No se sobrescribió nada. Regresando al menú principal.")
            return
        else:
            # Eliminar y recrear la carpeta
            for archivo in os.listdir(carpeta_destino):
                os.remove(os.path.join(carpeta_destino, archivo))
            os.rmdir(carpeta_destino)

    os.makedirs(carpeta_destino)

    # Crear imágenes personalizadas
    for idx, (participante, amigo) in enumerate(registros):
        # frase = frases_graciosas[random.randint(0, len(frases_graciosas) - 1)]  # Seleccionar frase aleatoria
        frase = seleccionar_frase_unica(frases_graciosas, frases_usadas)
        nombre_imagen = os.path.join(carpeta_destino, f"{idx + 1:02}Img.jpg")

        try:
            with Image.open(fondo_imagen) as img:
                draw = ImageDraw.Draw(img)
                ancho, alto = img.size

                # Calcular tamaño de la fuente proporcional al tamaño de la imagen y aumentar X puntos
                tamaño_fuente = int(min(ancho, alto) * 0.05) + 45
                try:
                    fuente = ImageFont.truetype("arial.ttf", tamaño_fuente)  # Verificar si existe la fuente
                except OSError:
                    print("No se encontró la fuente 'arial.ttf'. Usando la fuente predeterminada.")
                    fuente = ImageFont.load_default()

                # Texto a escribir
                texto = (
                    f"AMIGO SECRETO {anio_seleccionado}\n\n"
                    f"{participante}, tu amigo secreto es {amigo}\n"
                    f"{frase}"
                )

                # Calcular posición del texto centrado
                texto_bbox = draw.textbbox((0, 0), texto, font=fuente, anchor="mm")
                posicion = (ancho // 2, alto // 2)

                # Escribir texto centrado con negritas y borde blanco
                draw.multiline_text(
                    posicion, texto, font=fuente, fill="white",
                    stroke_width=2, stroke_fill="black", align="center", anchor="mm"
                )

                # Guardar la imagen
                try:
                    img.save(nombre_imagen)
                except Exception as e:
                    print(f"Hubo un error al grabar la imagen: {e}")

        except Exception as e:
            print(f"Hubo un error al abrir o procesar la imagen: {e}")

    print(f"Las imágenes del año {anio_seleccionado} se generaron correctamente en la carpeta '{carpeta_destino}'.")

# Menú principal
def menu():
    nombre_archivo = config.BaseDatos
    workbook = cargar_o_crear_excel(nombre_archivo)
    anio_actual = datetime.now().year
    fondo_imagen = config.ArchivoJPG

    while True:
        print("\n--- Menú Principal ---")
        print("1. Agregar o editar participantes")
        print("2. Listar participantes actual")
        print("3. Ver sorteos por año")
        print("4. Preguntar por participante")
        print("5. Realizar Sorteo", anio_actual)
        print("6. Ilustrar amigos por año")
        print("7. Salir")
        opcion = input("Elige una opción: ")

        if opcion == "1":
            agregar_o_editar_participantes(workbook, nombre_archivo)
        elif opcion == "2":
            listar_participantes(workbook)
        elif opcion == "3":
            try:
                anio = int(input("Ingresa el año que deseas consultar: "))
                ver_sorteos_por_anio(workbook, anio, anio_actual)
            except ValueError as e:
                print("Por favor, ingresa un año válido. Error: {e}")
        elif opcion == "4":  # Llamada al nuevo método
            preguntar_por_participante(workbook)
        elif opcion == "5":
            realizar_sorteo(workbook, anio_actual, nombre_archivo)
        elif opcion == "6":
            try:
                anio = int(input("Ingresa el año que deseas ilustrar: "))
                ilustrar_amigos_por_anio(workbook, anio, fondo_imagen)
            except ValueError as e:
                print("Por favor, ingresa un año válido. Error: {e}")
        elif opcion == "7":
            print("¡Hasta luego!")
            break
        else:
            print("Opción inválida. Por favor, intenta de nuevo.")

# Ejecutar el programa
if __name__ == "__main__":
    menu()